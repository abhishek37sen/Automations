import openpyxl

wk = openpyxl.load_workbook("G:\\Excel_Sheet_Arjun_RAwat.xlsx")
print(wk.sheetnames)
print("Active sheets"+wk.active.title)
sh = wk["Sheet1"]
# print(sh.title)
# print(sh["A3"].value)
# cl = sh.cell(3,1)
# print(cl.value)
# cl = sh.cell(column=3,row=1)
# print(cl.value)
# print(cl.row)
# print(cl.column)

# rows = sh.max_row
# columns = sh.max_column
# print(rows)
# print(columns)
# for i in range(1,rows+1):
#     for j in range(1,columns+1):
#         c =sh.cell(i,j)
#         print(c.value,"   ",end='')
#     print()

for r in sh['A1':'k14']:
    for c in r:
        print(c.value,"   ",end='')
    print()