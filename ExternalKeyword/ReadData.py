import openpyxl

def total_rows(sheetname):
    wk = openpyxl.load_workbook("G://Excel_Sheet_Arjun_RAwat.xlsx")
    sh = wk[sheetname]
    # print(sh.title)
    # print(type(sh.max_row))
    return sh.max_row
    # m_row = sh.make_row
    # return m_row
print(total_rows('Sheet1'))
def fetch_data(sheetname,row,col):
    wk = openpyxl.load_workbook("G://Excel_Sheet_Arjun_RAwat.xlsx")
    print(wk.sheetnames)
    # print(wk.active.title)
    sh = wk[sheetname]
    # print(type(row))
    cl = (sh.cell(int(row),int(col))).value
    # print(cl.value)
    return cl
# cl = fetch_data('Sheet1',1,1)
# print(cl)